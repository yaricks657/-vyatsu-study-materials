import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns


# Гистограмма для игрока и его статистики
def players_effectiveness(player, stat_for_x, stat_for_y, teams):
    wb = openpyxl.reader.excel.load_workbook('test.xlsx')
    # for sheets in wb.worksheets:
    ws = wb[player]
    i = 2
    index = []
    type(index)
    values = []
    type(values)
    team = []
    type(team)

    while ws[stat_for_x + str(i)].value is not None:
        index.append(str(ws[stat_for_x + str(i)].value))
        values.append(ws[stat_for_y + str(i)].value)
        team.append(str(ws[stat_for_x + str(i)].value) + str(ws[teams + str(i)].value))
        i = i + 1

    def addlabels(x, y):
        for ii in range(len(x)):
            plt.text(ii, y[ii], y[ii], horizontalalignment='center', verticalalignment='bottom',
                     fontdict={'fontweight': 500, 'size': 12})

    # making the bar chart on the data
    plt.bar(index, values)

    # calling the function to add value labels
    addlabels(index, values)

    # giving title to the plot
    plt.title("Average PPG through years for " + player)

    # giving X and Y labels
    plt.xlabel("Age")
    plt.ylabel("PPG")
    plt.xticks(index)
    plt.gca().set_xticklabels(index, rotation=60, horizontalalignment='right')

    # visualizing the plot
    plt.show()


# Диаграмма площади Unstacked для сравнения статистики десятилетий (PA)
def average_decade_stats(begin_of_the_first_decade, begin_of_the_second_decade, PA):
    first_data = []
    type(first_data)
    second_data = []
    type(second_data)
    years1 = []
    type(years1)
    years2 = []
    type(years2)
    for i in range(10):
        first_data.append(
            pd.read_excel(str(begin_of_the_first_decade + i) + '-' + str(begin_of_the_first_decade + i + 1) + '.xlsx'))
        second_data.append(pd.read_excel(
            str(begin_of_the_second_decade + i) + '-' + str(begin_of_the_second_decade + i + 1) + '.xlsx'))
        years1.append(str(begin_of_the_first_decade + i) + '-' + str(begin_of_the_first_decade + i + 1))
        years2.append(str(begin_of_the_second_decade + i) + '-' + str(begin_of_the_second_decade + i + 1))
    first_decade_years = np.array(years1)  # Массви первых лет
    second_decade_years = np.array(years2)  # Массив вторых лет
    x = []
    type(x)
    y1 = []
    type(y1)
    y2 = []
    type(y2)
    for i in range(10):
        y1.append(np.array(first_data[i][PA].values) * np.array(first_data[i]['G'].values))
        y2.append(np.array(second_data[i][PA].values) * np.array(second_data[i]['G'].values))
    main1 = []
    type(main1)
    main2 = []
    type(main2)
    for i in range(10):
        main1.append(np.sum(y1[i]))
        main2.append(np.sum(y2[i]))
    first_decade = np.array(main1).astype(int)  # Первый массив с данными за первое десятилетие
    second_decade = np.array(main2).astype(int)  # Второй массив с данными за второе десятилетие

    mycolors = ['tab:red', 'tab:blue', 'tab:green', 'tab:orange', 'tab:brown', 'tab:grey', 'tab:pink', 'tab:olive']
    columns = ['2010-2020 years', '2000-2010 years']

    # Draw Plot
    fig, ax = plt.subplots(1, 1, figsize=(16, 9), dpi=80)
    ax.fill_between(second_decade_years, y1=second_decade, y2=0, label=columns[0], alpha=0.5, color=mycolors[0],
                    linewidth=2)
    ax.fill_between(second_decade_years, y1=first_decade, y2=0, label=columns[1], alpha=0.5, color=mycolors[1],
                    linewidth=2)

    # Decorations
    ax.set_title('Evolution of ' + PA + ' basketball. ' + PA + ' per year', fontsize=18)
    ax.set(ylim=[0, 200000])
    ax.legend(loc='best', fontsize=12)
    plt.xlabel("Years")
    plt.ylabel("3PA")

    # Draw Tick lines
    for y in np.arange(2.5, 30.0, 2.5):
        plt.hlines(y, xmin=0, xmax=len(x), colors='black', alpha=0.3, linestyles="--", lw=0.5)

    # Lighten borders
    plt.gca().spines["top"].set_alpha(0)
    plt.gca().spines["bottom"].set_alpha(.3)
    plt.gca().spines["right"].set_alpha(0)
    plt.gca().spines["left"].set_alpha(.3)
    plt.show()


# Boxplot, показывающий среднее количество чего-то по возрасту
def average_pts_stats(begin_of_the_first_decade, number_of_years, stat):
    first_data = []
    type(first_data)
    for i in range(number_of_years):  # Анализ возраста за 5 лет
        first_data.append(
            pd.read_excel(str(begin_of_the_first_decade + i) + '-' + str(begin_of_the_first_decade + i + 1) + '.xlsx'))
    y1 = []
    type(y1)
    y2 = []
    type(y2)
    for i in range(number_of_years):
        y1.append(np.array(first_data[i]['Age'].values).astype(int))  # Первый массив с данными о возрасте
        y2.append(np.array(first_data[i][stat].values))  # Второй массив со статистикой
    ages = np.concatenate(y1, 0)  # Массив со всеми возрастами
    statistics = np.concatenate(y2, 0)  # Массив со всей статистикой

    age19 = []
    type(age19)
    age20 = []
    type(age20)
    age21 = []
    type(age21)
    age22 = []
    type(age22)
    age23 = []
    type(age23)
    age24 = []
    type(age24)
    age25 = []
    type(age25)
    age26 = []
    type(age26)
    age27 = []
    type(age27)
    age28 = []
    type(age28)
    age29 = []
    type(age29)
    age30 = []
    type(age30)
    age31 = []
    type(age31)
    age32 = []
    type(age32)
    age33 = []
    type(age33)
    age34 = []
    type(age34)
    age35 = []
    type(age35)
    age36 = []
    type(age36)
    age37 = []
    type(age37)
    age38 = []
    type(age38)
    age39 = []
    type(age39)
    age40 = []
    type(age40)
    age41 = []
    type(age41)
    age42 = []
    type(age42)
    age43 = []
    type(age43)

    for i in range(len(ages)):
        if ages[i] == 19:
            age19.append(statistics[i])
        if ages[i] == 20:
            age20.append(statistics[i])
        if ages[i] == 21:
            age21.append(statistics[i])
        if ages[i] == 22:
            age22.append(statistics[i])
        if ages[i] == 23:
            age23.append(statistics[i])
        if ages[i] == 24:
            age24.append(statistics[i])
        if ages[i] == 25:
            age25.append(statistics[i])
        if ages[i] == 26:
            age26.append(statistics[i])
        if ages[i] == 27:
            age27.append(statistics[i])
        if ages[i] == 28:
            age28.append(statistics[i])
        if ages[i] == 29:
            age29.append(statistics[i])
        if ages[i] == 30:
            age30.append(statistics[i])
        if ages[i] == 31:
            age31.append(statistics[i])
        if ages[i] == 32:
            age32.append(statistics[i])
        if ages[i] == 33:
            age33.append(statistics[i])
        if ages[i] == 34:
            age34.append(statistics[i])
        if ages[i] == 35:
            age35.append(statistics[i])
        if ages[i] == 36:
            age36.append(statistics[i])
        if ages[i] == 37:
            age37.append(statistics[i])
        if ages[i] == 38:
            age38.append(statistics[i])
        if ages[i] == 39:
            age39.append(statistics[i])
        if ages[i] == 40:
            age40.append(statistics[i])
        if ages[i] == 41:
            age41.append(statistics[i])
        if ages[i] == 42:
            age42.append(statistics[i])
        if ages[i] == 43:
            age43.append(statistics[i])

    data = [age19, age20, age21, age22, age23, age24, age25, age26, age27, age28, age29, age30, age31, age32, age33,
            age34, age35, age36, age37, age38, age39, age40, age41, age42, age43]

    # Creating axes instance
    fig = plt.figure(figsize=(10, 7))
    ax = fig.add_subplot(111)

    # Creating plot
    bp = ax.boxplot(data, vert=1)

    # x-axis labels
    ax.set_xticklabels(
        ['19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36',
         '37', '38', '39', '40', '41', '42', '43'])
    plt.xlabel("Age")
    plt.ylabel(stat)

    # Adding title
    plt.title(stat + ' per age for players in ' + str(begin_of_the_first_decade) + '-' + str(
        begin_of_the_first_decade + number_of_years))

    # show plot
    plt.show()

#Матрица корреляции для зависимостей
def matrix_correlogram(player):
    df = pd.read_excel('test.xlsx',sheet_name=player)
    # Plot
    plt.figure(figsize=(12, 10), dpi=80)
    sns.heatmap(df.corr(), xticklabels=df.corr().columns, yticklabels=df.corr().columns, cmap='RdYlGn', center=0,
                annot=True)

    # Decorations
    plt.title(player, fontsize=22)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.show()


# Точечный график для позиций
def player_position_efficient(year, stat):
    df = pd.read_excel(str(year) + '-' + str(year + 1) + '.xlsx')
    # Draw Plot
    pg_sg = []
    type(pg_sg)
    sf_pf = []
    type(sf_pf)
    c = []
    type(c)
    pg_sg_time = []
    type(pg_sg_time)
    sf_pf_time = []
    type(sf_pf_time)
    c_time = []
    type(c_time)
    for i in range(len(df["Age"])):
        if df['Pos'][i] in ['PG', 'SG', 'PG-SG', 'SG-PG'] and df['MP'][i] >= 5:
            pg_sg.append(df[stat][i])
            pg_sg_time.append(df['MP'][i])
        if df['Pos'][i] in ['SF', 'PF', 'SG-SF', 'SF-SG', 'SF-PF', 'PF-SF'] and df['MP'][i] >= 5:
            sf_pf.append(df[stat][i])
            sf_pf_time.append(df['MP'][i])
        if df['Pos'][i] in ['C', 'C-PF', 'PF-C'] and df['MP'][i] >= 5:
            c.append(df[stat][i])
            c_time.append(df['MP'][i])
    plt.figure(figsize=(16, 10), dpi=80)
    plt.scatter(c_time, c, color='tab:blue',s=120,data=1, label="Centers")
    plt.scatter(sf_pf_time, sf_pf, color='tab:green', s=60, label="Forwards")
    plt.scatter(pg_sg_time, pg_sg, color='tab:red', s=30, label="Guards")

    # Decoration
    plt.xticks(rotation=0, fontsize=12, horizontalalignment='center',
               alpha=.7)
    plt.yticks(fontsize=12, alpha=.7)
    plt.title(stat + " among guards, forwards and centers", fontsize=22)
    plt.grid(axis='both', alpha=.3)
    plt.xlabel("Minutes per game")
    plt.ylabel(stat)
    plt.legend()
    # Remove borders
    plt.gca().spines["top"].set_alpha(0.0)
    plt.gca().spines["bottom"].set_alpha(0.3)
    plt.gca().spines["right"].set_alpha(0.0)
    plt.gca().spines["left"].set_alpha(0.3)
    plt.show()