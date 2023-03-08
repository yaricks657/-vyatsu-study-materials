import openpyxl
from openpyxl import Workbook

final = Workbook()

name = 1946

for i in range(76):#76
    date = str(name + i) + '-' + str(name + 1 + i)
    book = openpyxl.reader.excel.load_workbook(date + '.xlsx')
    book.active = 0
    ws1 = book.active
    counter = 2
    player = book.active['B' + str(counter)].value
    print('Уже ' + str(i+1) + '-й файл')
    final.save('test.xlsx')
    while player is not None:
        if player[len(player) - 1] == '*':
            player = player[:len(player) - 1]
        if player not in final.sheetnames:
            ws2 = final.create_sheet(player, counter - 2)
            ws2['A1'] = 'Years'
            ws2['B1'] = 'Player'
            ws2['C1'] = 'Pos'
            ws2['D1'] = 'Age'
            ws2['E1'] = 'Team'
            ws2['F1'] = 'GP'
            ws2['G1'] = 'GS'
            ws2['H1'] = 'MP'
            ws2['I1'] = 'FG'
            ws2['J1'] = 'FGA'
            ws2['K1'] = 'FG%'
            ws2['L1'] = '3P'
            ws2['M1'] = '3PA'
            ws2['N1'] = '3P%'
            ws2['O1'] = '2P'
            ws2['P1'] = '2PA'
            ws2['Q1'] = '2P%'
            ws2['R1'] = 'eFG%'
            ws2['S1'] = 'FT'
            ws2['T1'] = 'FTA'
            ws2['U1'] = 'FT%'
            ws2['V1'] = 'ORB'
            ws2['W1'] = 'DRB'
            ws2['X1'] = 'TRB'
            ws2['Y1'] = 'AST'
            ws2['Z1'] = 'STL'
            ws2['AA1'] = 'BLK'
            ws2['AB1'] = 'TOV'
            ws2['AC1'] = 'PF'
            ws2['AD1'] = 'PTS'
        else:
            ws2 = final[player]
        for j in range(2, 608):
            if ws2['A' + str(j)].value is None:
                ws2['A' + str(j)] = date
                ws2['B' + str(j)] = player
                ws2['C' + str(j)] = ws1['C' + str(counter)].value
                ws2['D' + str(j)] = ws1['D' + str(counter)].value
                ws2['E' + str(j)] = ws1['E' + str(counter)].value
                ws2['F' + str(j)] = ws1['F' + str(counter)].value
                ws2['G' + str(j)] = ws1['G' + str(counter)].value
                ws2['H' + str(j)] = ws1['H' + str(counter)].value
                ws2['I' + str(j)] = ws1['I' + str(counter)].value
                ws2['J' + str(j)] = ws1['J' + str(counter)].value
                ws2['K' + str(j)] = ws1['K' + str(counter)].value
                ws2['L' + str(j)] = ws1['L' + str(counter)].value
                ws2['M' + str(j)] = ws1['M' + str(counter)].value
                ws2['N' + str(j)] = ws1['N' + str(counter)].value
                ws2['O' + str(j)] = ws1['O' + str(counter)].value
                ws2['P' + str(j)] = ws1['P' + str(counter)].value
                ws2['Q' + str(j)] = ws1['Q' + str(counter)].value
                ws2['R' + str(j)] = ws1['R' + str(counter)].value
                ws2['S' + str(j)] = ws1['S' + str(counter)].value
                ws2['T' + str(j)] = ws1['T' + str(counter)].value
                ws2['U' + str(j)] = ws1['U' + str(counter)].value
                ws2['V' + str(j)] = ws1['V' + str(counter)].value
                ws2['W' + str(j)] = ws1['W' + str(counter)].value
                ws2['X' + str(j)] = ws1['X' + str(counter)].value
                ws2['Y' + str(j)] = ws1['Y' + str(counter)].value
                ws2['Z' + str(j)] = ws1['Z' + str(counter)].value
                ws2['AA' + str(j)] = ws1['AA' + str(counter)].value
                ws2['AB' + str(j)] = ws1['AB' + str(counter)].value
                ws2['AC' + str(j)] = ws1['AC' + str(counter)].value
                ws2['AD' + str(j)] = ws1['AD' + str(counter)].value
                #final.save('test.xlsx')
                break
        #print(counter)
        counter = counter + 1
        player = book.active['B' + str(counter)].value

final.save('test.xlsx')
